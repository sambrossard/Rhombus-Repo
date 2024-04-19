import re
import os
import pprint
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from openpyxl import Workbook


class RFPDataExtractor:
    def scan_pdf(self, file_path):
        try:
            with open(file_path, 'rb') as file:
                reader = PdfReader(file)
                text = ''
                for page in reader.pages:
                    text += page.extract_text() + '\n'  
            return text
        except Exception as e:
            print("Error:", e)
            return None

    def structure_data(self, text):
        if not text:
            return None
        
        data = {}
        qa_pairs = re.findall(r'Q:(.*?)A:(.*?)(?=Q:|$)', text, re.DOTALL)
        for pair in qa_pairs:
            question = pair[0].strip()
            answer = pair[1].strip()
            data[question] = answer
        pprint.pprint(data)
        return data
    
class ExcelDataExtractor:
    def scan_excel(self, file_path):
        try:
            wb = load_workbook(file_path)
            sheet = wb.active
            text = ''
            for row in sheet.iter_rows(values_only=True):
                text += ' '.join([str(cell) for cell in row]) + '\n'
            return text
        except Exception as e:
            print("Error:", e)
            return None

    def structure_excel_data(self, text):
        if not text:
            return None
        
        data = {}
        excel_qa_pairs = re.findall(r'Q:(.*?)A:(.*?)(?=Q:|$)', text, re.DOTALL)
        for pair in excel_qa_pairs:
            question = pair[0].strip()
            answer = pair[1].strip()
            data[question] = answer
        pprint.pprint(data)
        return data

class DataExtractorCaller:
    def __init__(self):
        self.extractor = RFPDataExtractor()
        self.extractor2 = ExcelDataExtractor()

    def extract_pdf_data(self, file_path):
        text = self.extractor.scan_pdf(file_path)
        if text:
            structured_data = self.extractor.structure_data(text)
            return structured_data
        else:
            print("Failed to extract text from the PDF.")
            return None
    def extract_excel_data(self, file_path):
        text = self.extractor2.scan_excel(file_path)
        if text:
            structured_data = self.extractor2.structure_excel_data(text)
            return structured_data
        else:
            print("Failed to extract text from the Excel.")
            return None

if __name__ == "__main__":
    caller = DataExtractorCaller()
    structured_data = caller.extract_pdf_data('/Users/samuelbrossard/python/projectvenv/RFP_Test_pdf.pdf')
    structured_excel_data = caller.extract_excel_data('/Users/samuelbrossard/python/projectvenv/Test_RFP_Excel.xlsx')
    if structured_data:
        print("Data extraction successful.")
    else:
        print("Data extraction failed.")




